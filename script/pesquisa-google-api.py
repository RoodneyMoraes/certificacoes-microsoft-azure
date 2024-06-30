import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

# Função para pesquisar no Google API
def pesquisar_google_api(query_pesquisa):
    url_da_api = "https://www.googleapis.com/customsearch/v1"
    parametros = {
        "key": "AIzaSyBR8WH6whkKYXYF8uaIdBjqKV0hYdUClHE",
        "cx": "66a59a5d64de449a9",
        "q": query_pesquisa
    }
    
    try:
        response = requests.get(url_da_api, params=parametros)
        response.raise_for_status()  # Força uma exceção se o status não for 200
        resultados = response.json().get("items", [])
        if resultados:
            titulo = resultados[0].get("title", "Sem título")
            link = resultados[0].get("link", "Sem link")
        else:
            titulo = "Nenhum resultado encontrado."
            link = ""
    except Exception as e:
        titulo = f"Erro: {str(e)}"
        link = ""
    return titulo, link

# Função para salvar os resultados em um arquivo Excel
def salvar_em_excel(resultados, nome_arquivo="resultados.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Título", "Link"])  # Adiciona cabeçalho
    for titulo, link in resultados:
        if link:
            ws.append([titulo, link])
            cell_link = ws.cell(row=ws.max_row, column=2)
            if not titulo.startswith("Erro:"):
                cell_link.hyperlink = link
                cell_link.font = Font(color="0000FF", underline="single")
        else:
            ws.append([titulo, "Sem link"])
    wb.save(nome_arquivo)
    print(f"Arquivo '{nome_arquivo}' salvo com sucesso.")

# Execução principal
if __name__ == "__main__":
    resultados = []

    for i in range(1, 105):
        query_pesquisa = f"examtopics az-104 topic 2 question {i}"
        titulo, link = pesquisar_google_api(query_pesquisa)
        resultados.append((titulo, link))
        print(f"Processado: questão {i}.")
        time.sleep(1)  # Delay de 1 segundo para evitar erro 429

    salvar_em_excel(resultados)